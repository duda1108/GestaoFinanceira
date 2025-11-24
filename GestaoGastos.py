from openpyxl import Workbook
from GestaoFinanceira.Transacoes import transacoes
from GestaoFinanceira.ListarTransacoes import vizualizar
import os
from openpyxl import Workbook, load_workbook
#guardará os cadastros
import json
import hashlib

#====================== Criaçao de cadastro, e verificaçao de Login =========================
#============================================================================================
USERS_FILE = "users.json"

#--------------tranforma qualquer texto em uma sequencia fixa de caracteres-----------------
def _hash_password(senha: str) -> str:
    return hashlib.sha256(senha.encode("utf-8")).hexdigest()


#-----------verifica os user no json e verifica se existe --------------
def load_users(filename=USERS_FILE):
    try:
        with open(filename, "r", encoding="utf-8") as f:
            data = json.load(f)
            # data vai ser tipo: {usuario: hashed_password}
            return data if isinstance(data, dict) else {}
    except FileNotFoundError:
        return {}
    except Exception as e:
        print("Error ao carregar usuario no .json:", e)
        return {}

#-----------------salva novos usuarios do json------------------
def save_users(usuarios, filename=USERS_FILE):
    try:
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(usuarios, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print("Erro ao salvar usuario no .json:", e)

usuarios = load_users()


#=======================CADASTRO==========================
#=========================================================
def cadastrar_usuario():
    print("--------- CADASTRO ---------")
    nome = input("Digite um nome de usuário: ").strip()
    if nome == "":
        print("Nome de usuário não pode ser vazio.")
        return
    if nome in usuarios:
        print("Usuário já existe. Tente outro nome.")
        return

    senha = input("Digite uma senha: ").strip()
    if senha == "":
        print("Senha não pode ser vazia.")
        return

    #--------- armazena a senha "hash" no pc --------------
    usuarios[nome] = _hash_password(senha)
    try:
        save_users(usuarios)
        print(f"Usuário '{nome}' cadastrado com sucesso!")
    except Exception as e:
        print("Erro ao salvar usuário:", e)



#=====================LOGIN======================
#================================================
def fazer_login():
    print("\n--------- LOGIN ---------")
    nome = input("Usuário: ").strip()
    senha = input("Senha: ").strip()

    #hash password é para guardar a senha
    if nome in usuarios and usuarios[nome] == _hash_password(senha):
        print(f"Login bem-sucedido! Bem-vindo, {nome}.")
        return nome
    else:
        print("Usuário ou senha inválidos.")
        return None



#=======================CRIAÇÃO PLANILHA=======================
#=============================================================

FILENAME = "Gestao_Financeira.xlsx"
def criar_planilha(filename=FILENAME):

#-----------Verficação se planilha ja existe, para que--------------------
#---------nao crie outra e sobreponha á uma ja existente------------------
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "Janeiro"
    meses = ["Fevereiro", "Março", "Abril", "Maio", "Junho",
             "Julho", "Agosto", "Setembro", "Outubro",
             "Novembro", "Dezembro"]
    for m in meses:
        wb.create_sheet(m)

    for sheet in wb.worksheets:
        sheet.append(["ID", "Tipo", "Categoria", "Valor", "Data", "Descrição"])

    wb.save(filename)
    return wb, ws



#================================MENU INICIAL==================================
#==================================TKINTER=====================================
import tkinter as tk
from tkinter import messagebox
import sys
import builtins
from tkinter import simpledialog


#Interface inicial
janela = tk.Tk()
janela.title("Gestão Financeira")
janela.geometry("400x300")

#--------------------------------Seções de menu, login e cadastro--------------------------------------
frame_menu = tk.Frame(janela)
frame_login = tk.Frame(janela)
frame_cadastro = tk.Frame(janela)
frame_principal = tk.Frame(janela)
for f in (frame_menu, frame_login, frame_cadastro, frame_principal):
    f.place(relx=0, rely=0, relwidth=1, relheight=1)

def mostrar(frame):
    frame.tkraise()


tk.Label(frame_menu, text="MENU INICIAL", font=("Arial",14)).pack(pady=10)
#Botões pra adicionar ao menu de login
tk.Button(frame_menu, text="Cadastrar usuário", width=20, command=lambda: mostrar(frame_cadastro)).pack(pady=5)
tk.Button(frame_menu, text="Login", width=20, command=lambda: mostrar(frame_login)).pack(pady=5)
tk.Button(frame_menu, text="Sair", width=20, command=janela.destroy).pack(pady=5)


#-------------------------------------- CADASTRO ------------------------------------------
tk.Label(frame_cadastro, text="Cadastro", font=("Arial",14)).pack(pady=10)
cad_nome = tk.Entry(frame_cadastro); cad_senha = tk.Entry(frame_cadastro, show="*")
cad_nome.pack(pady=5); cad_senha.pack(pady=5)
def cadastrar_ui():
    nome = cad_nome.get().strip(); senha = cad_senha.get().strip()
    if not nome or not senha:
        messagebox.showerror("Erro","Preencha todos os campos antes, tente novamente"); return
    if nome in usuarios:
        messagebox.showerror("Erro","Usuário já existe, escolha outro."); return
    usuarios[nome] = _hash_password(senha); save_users(usuarios)
    messagebox.showinfo("Sucesso","Cadastro realizado com sucesso"); mostrar(frame_menu)
tk.Button(frame_cadastro, text="Cadastrar", command=cadastrar_ui).pack(pady=5)
tk.Button(frame_cadastro, text="Voltar", command=lambda: mostrar(frame_menu)).pack()


#---------------------------------------- LOGIN -------------------------------------------------
tk.Label(frame_login, text="Login", font=("Arial",14)).pack(pady=10)
login_nome = tk.Entry(frame_login); login_senha = tk.Entry(frame_login, show="*")
login_nome.pack(pady=5); login_senha.pack(pady=5)
usuario_logado = None

#Função para buscar o login
def login_ui():
    global usuario_logado
    nome = login_nome.get().strip(); senha = login_senha.get().strip()
    if nome in usuarios and usuarios[nome] == _hash_password(senha):
        usuario_logado = nome
        messagebox.showinfo("Sucesso", f"Bem-vindo, {nome}!")
        mostrar(frame_principal)
    else:
        messagebox.showerror("Erro","Usuário ou senha inválidos.")
tk.Button(frame_login, text="Entrar", command=login_ui).pack(pady=5)
tk.Button(frame_login, text="Voltar", command=lambda: mostrar(frame_menu)).pack()

#Interface da Gestao Financeira
tk.Label(frame_principal, text="Menu Financeiro", font=("Arial",14)).pack(pady=10)
wb, ws = criar_planilha(FILENAME)

#Função para executar ações apos a escolha de um dos numeros do menu da gestao finaceira
def executar(op):

    #Criará uma janela apos a escolha da ação, para preencher com os dados necessarios
    out_win = tk.Toplevel(janela)
    out_win.title("Saída / Interação")
    out_win.geometry("600x400")

    txt = tk.Text(out_win, wrap="word")
    txt.pack(expand=True, fill="both")

    #Botoes para limpar o texto ou fechas as janelas
    btn_frame = tk.Frame(out_win)
    btn_frame.pack(fill="x", pady=4)
    tk.Button(btn_frame, text="Limpar", command=lambda: txt.delete("1.0", "end")).pack(side="left", padx=6)
    tk.Button(btn_frame, text="Fechar", command=out_win.destroy).pack(side="right", padx=6)

    #Peça Chave do tkinter
    #TUDO o que sera printado no terminal, agora cai no tkinter
    def insira_o_texto(s):
        try:
            txt.insert("end", s)  #escreve em formato string
            txt.see("end")              #rola a janela para mostrar o final
            out_win.update_idletasks()  #mantem a janela enquanto escreve
        except Exception:
            pass

    #Função que "engana" o tkinter, para que tudo que voce escreva no tkinter, ele envia pela função "insira o texto"
    class _StdRedirect:
        def write(self, s):
            insira_o_texto(s)
        def flush(self):
            pass

    #Salva valores originais antes e redirecionar
    old_stdout = sys.stdout
    old_stderr = sys.stderr
    old_input = builtins.input

    #Qualquer print vira "insir_o_texto"
    sys.stdout = _StdRedirect()
    sys.stderr = _StdRedirect()

    #Toda vez que o codigo pede algum input, aparecerá na janela, de forma bonita
    def input_consumidor(prompt=""):
        # mostra diálogo para o usuário digitar
        res = simpledialog.askstring("Entrada", prompt or "Digite:", parent=out_win)
        return "" if res is None else res

    builtins.input = input_consumidor

    # garantir que a janela apareça
    out_win.update()
    out_win.lift()
    out_win.focus_force()

    try:
        # executar a operação
        if op in (3, 4, 5):
            vizualizar(op, ws, wb)
        else:
            transacoes(op, ws, wb)

        try:
            wb.save(FILENAME)
        except Exception as e:
            #Caso nao consiga salvar, mostre mensagem curta e registre no Text
            insira_o_texto(f"\nErro ao salvar planilha: {e}\n")
            messagebox.showwarning("Aviso", "Não foi possível salvar a planilha.")
    except Exception as e:
        # mostrar erro curto ao usuário e log completo no Text
        import traceback
        tb = traceback.format_exc()
        insira_o_texto("\n=== Ocorreu um erro ===\n")
        insira_o_texto(tb + "\n")
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}\nVeja na janela para mais detalhes.")
    finally:
        #Salva valores originais antes e redirecionar denovo
        sys.stdout = old_stdout
        sys.stderr = old_stderr
        builtins.input = old_input

    wb.save(FILENAME)

#Menu com as opçoes de ações existentes
ops = [("1 - Adicionar transação",1),("2 - Remover transação",2),
       ("3 - Listar transações por categoria",3),("4 - Listar transações por período",4),
       ("5 - Saldo acumulado ao longo do tempo",5), ("6 - Saldo do periodo",6), ("0 - Encerrar sessão",0)]
for txt, cod in ops:
    tk.Button(frame_principal, text=txt, command=lambda c=cod: executar(c)).pack(pady=2)
tk.Button(frame_principal, text="Logout", command=lambda: mostrar(frame_menu)).pack(pady=10)

mostrar(frame_menu)
janela.mainloop()
