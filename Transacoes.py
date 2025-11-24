# GestaoFinanceira/Transacoes.py
from openpyxl import load_workbook
import os
from datetime import datetime

FNAME = "Gestao_Financeira.xlsx"
MESES = ["Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
CATEGORIAS = {"1":"Alimentação","2":"Moradia","3":"Vestuário","4":"Outros"}


#=============================== ID ====================================
def ID(wb):
    max_id = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            try:
                cell = row[0]
                if cell is None:
                    continue
                iid = int(cell)
                if iid > max_id:
                    max_id = iid
            except Exception:
                continue
    return max_id + 1

#================================= Transasções =================================
def transacoes(n, ws, wb):
    print(f"[DEBUG] transacoes() chamado com n={n}")

    #-------------------------- ADICIONAR -----------------------------
    if n == 1:
        #--------------------------------------- DATA ------------------------------------------------
        print("-----------------------------------------------------------------------")
        data_str = input("Informe a data da transação (ex: 12/10 ou 12/10/2025): ").strip()
        if data_str == "":
            print("Data não pode ser vazia. Operação abortada.")
            return None


        #-------------------------------------- TIPO -------------------------------------------
        try:
            tipo_in = int(input("Me informe qual foi o tipo de transação:\n1. Entrada \n2. Saída\n"))
        except ValueError:
            print("Entrada inválida. Operação abortada.")
            return None
        if tipo_in == 1:
            tipo = "Entrada"
        elif tipo_in == 2:
            tipo = "Saída"
        else:
            print("Você deve digitar 1 ou 2. Operação abortada.")
            return None


        #------------------------------------ CATEGORIA -----------------------------------------
        categoria = input("Qual seria a categoria da transação:\n1- Alimentação\n2- Moradia\n3- Vestuário\n4- Outros\nEscolha (número ou nome): ").strip()
        if categoria == "":
            print("Categoria vazia. Operação abortada.")
            return None
        if categoria in CATEGORIAS:
            categoria = CATEGORIAS[categoria]
        else:
            cap = categoria.capitalize()
            if cap in CATEGORIAS.values():
                categoria = cap
            else:
                print("Categoria inválida. Operação abortada.")
                return None


        #------------------------------- VALOR ------------------------------------
        val_in = input("Informe o valor (use . como separador decimal): ").strip()
        try:
            valor = float(val_in)
            if valor <= 0:
                print("Valor deve ser positivo. Operação abortada.")
                return None
        except ValueError:
            print("Valor inválido. Operação nao concluída.")
            return None

        descricao = input("Descrição (opcional): ").strip()


        #------------------------------- ESCOLHA DE PLANILHA ---------------------------------------
        escolha = input("Deseja usar a planilha ativa (que seria janeiro) 'a', ou escolher outro mês 'c'? Digite 'a' ou 'c': ").strip().lower()
        if escolha == 'c':
            meses_full = ["Janeiro"] + MESES
            print("Escolha o mês (1-12):")
            for idx, m in enumerate(meses_full, start=1):
                print(f"{idx:2d} - {m}")
            sel_in = input("Mês (1-12): ").strip()
            try:
                sel = int(sel_in)
            except ValueError:
                print("Entrada inválida!")
                return None
            if 1 <= sel <= 12:
                sheet_name = meses_full[sel-1]
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
                ws_target = wb[sheet_name]
            else:
                print("Número fora das opções!")
                return None
        else:
            ws_target = ws

        tid = ID(wb)
        ws_target.append([tid, tipo, categoria, valor, data_str, descricao])
        wb.save(FNAME)
        print(f"Transação adicionada com ID {tid} na planilha {ws_target.title}.")
        return None

#======================= Remover transação =========================
    elif n == 2:
        # --- Lista todas as transações antes de pedir o ID ---
        print("\n========== Lista de todas as transações ==========\n")
        encontrou_algo = False
        for wsx in wb.worksheets:
            any_in_sheet = False
            for row in wsx.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                # garante que a tupla tenha pelo menos 6 elementos
                row = tuple(row) + (None,) * max(0, 6 - len(row))
                tid, tipo, categoria, valor, data, desc = row
                if tid is None:
                    continue
                if not any_in_sheet:
                    print(f"--- Planilha: {wsx.title} ---")
                    any_in_sheet = True
                valor_str = f"R${valor}" if valor is not None else "R$0"
                print(f"ID:{tid} | {tipo} | {categoria} | {valor_str} | {data} | {desc}")
                encontrou_algo = True

        if not encontrou_algo:
            print("Nenhuma transação encontrada.\n")

        id_in = input("Informe o ID da transação a remover (ou 'c' para cancelar): ").strip()
        if id_in.lower() == 'c' or id_in == "":
            print("Operação cancelada.")
            return None
        try:
            id_rem = int(id_in)
        except ValueError:
            print("ID inválido.")
            return None


        #Apos a seleção do ID, busca em qual planilha pertence e solicita confirmação para poder remover
        # Apos a seleção do ID, busca em qual planilha pertence e solicita confirmação para poder remover
        for wsx in wb.worksheets:
            # usar values_only=True para trabalhar só com valores e evitar .value
            for row_idx, row in enumerate(wsx.iter_rows(min_row=2, values_only=True), start=2):
                if not row:
                    continue

                # garante que a tupla tenha pelo menos 6 elementos
                row = tuple(row) + (None,) * max(0, 6 - len(row))
                cell_id = row[0]

                # se a célula de ID for None, pula
                if cell_id is None:
                    continue

                if cell_id == id_rem:
                    tipo, categoria, valor, data, desc = row[1], row[2], row[3], row[4], row[5]
                    print(
                        f"Encontrado na planilha {wsx.title}: ID {cell_id} | {tipo} | {categoria} | {valor} | {data} | {desc}")
                    confirma = input("Confirmar remoção? (s/n): ").strip().lower()
                    if confirma == 's':
                        wsx.delete_rows(row_idx, 1)
                        wb.save(FNAME)
                        print("Transação removida.")
                        return None
                    else:
                        print("Remoção cancelada.")
                        return None
        # Caso tenha escolhido um ID que nao existe
        print("ID não encontrado.")
        return None

    #======================== Encerrar ===========================
    elif n == 0:
        print("Encerrando...")
        return "quebra"

    #outras opções: retorna e ua vez para evitar loop
    else:
        return None
